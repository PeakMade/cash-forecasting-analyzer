import pandas as pd
import openpyxl

file_path = 'sample_files/155 River Oaks Cash Forecast - 10.2025.xlsx'

# Check with openpyxl
print("="*80)
print("CHECKING WITH OPENPYXL")
print("="*80)
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f'Max column: {ws.max_column}')
print(f'Max row: {ws.max_row}')

print(f'\nChecking for hidden columns:')
hidden_cols = []
for col in range(1, min(100, ws.max_column+1)):
    col_letter = openpyxl.utils.get_column_letter(col)
    is_hidden = ws.column_dimensions[col_letter].hidden
    if is_hidden:
        hidden_cols.append(f'{col}({col_letter})')
    if col <= 25:
        print(f'Column {col:3d} ({col_letter:3s}): hidden={is_hidden}')

print(f'\nHidden columns: {", ".join(hidden_cols) if hidden_cols else "None"}')

# Now read with pandas to see what we get
print("\n" + "="*80)
print("READING WITH PANDAS")
print("="*80)
df = pd.read_excel(file_path, sheet_name=0, header=None)
print(f'Pandas DataFrame shape: {df.shape}')

# Try to find 2025 data
print("\nSearching for '2025' in the dataframe:")
for row_idx in range(min(15, len(df))):
    for col_idx in range(len(df.columns)):
        cell = df.iloc[row_idx, col_idx]
        if cell and '2025' in str(cell):
            print(f'Found "2025" at row {row_idx}, col {col_idx}: {cell}')

print("\nLooking for 'Actual' or 'Budget' status indicators:")
for row_idx in range(min(15, len(df))):
    row_vals = df.iloc[row_idx, :].tolist()
    if any('actual' in str(v).lower() or 'budget' in str(v).lower() for v in row_vals if v):
        print(f'Row {row_idx}: {row_vals[:15]}')

print("\nRow 6-8 content (first 15 cols):")
for i in [6, 7, 8]:
    if i < len(df):
        print(f'Row {i}: {df.iloc[i, :15].tolist()}')
